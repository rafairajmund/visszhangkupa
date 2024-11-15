from flask import Flask, g, render_template, request, redirect, session, url_for, jsonify
from unidecode import unidecode
import openpyxl
from flask_babel import Babel, _

app = Flask(__name__)
app.secret_key = 'fasz'
EXCEL_FILE_PATH = r'D:\visszhang\EREDMENYEK.xlsx'

# Konfiguráljuk a Babelt
app.config['LANGUAGES'] = ['hu', 'ro']
app.config['BABEL_DEFAULT_LOCALE'] = 'hu'
app.config['BABEL_DEFAULT_TIMEZONE'] = 'UTC'
babel = Babel()

@app.before_request
def before_request():
    lang = session.get('lang', None)
    if lang:
        g.lang = lang
    else:
        g.lang = app.config['BABEL_DEFAULT_LOCALE']
# Nyelv kiválasztása a session alapján
def get_locale():
    return session.get('lang', app.config['BABEL_DEFAULT_LOCALE'])
# Babel inicializálása az alkalmazással
babel.init_app(app, locale_selector=get_locale)
@app.route('/')
def index():
    players = get_players()
    return render_template('index.html', players=players)


@app.route('/change_language', methods=['POST'])
def change_language():
    lang = request.json.get('lang')  # Ha az AJAX JSON formában küldi az adatot
    if lang:
        session['lang'] = lang
    return jsonify(success=True)  # Változtasd meg, ha szükséges
@app.route('/process', methods=['POST'])
def process():
    # A process függvény többi része változatlan marad
    team1_name = request.form['team1_name']
    team2_name = request.form['team2_name']
    team1_score = int(request.form['team1_score'])
    team2_score = int(request.form['team2_score'])
    group = request.form['group']

    # Update match results in the Excel file
    update_excel(EXCEL_FILE_PATH, group, team1_name, team1_score, team2_name, team2_score)

    # Handle yellow-carded players
    yellow_card_players = request.form.getlist('yellow_card_players[]')
    for player in yellow_card_players:
        if player:
            update_card(player, 'FFFF00')  # Add yellow color

    # Handle red-carded players
    red_card_players = request.form.getlist('red_card_players[]')
    for player in red_card_players:
        if player:
            update_card(player, 'FF0000')  # Add red color

    return redirect(url_for('index'))

@app.route('/search_player', methods=['POST'])
def search_player():
    query = unidecode(request.json.get('query', '').lower().strip())
    if not query:
        return jsonify([])

    players = get_players()
    results = [p['name'] for p in players if query in unidecode(p['name'].lower())]
    return jsonify(results)

def update_card(player_name, color):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook['Sarga-piros lapok']

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if unidecode(row[0].value.lower()) == unidecode(player_name.lower()):
                col = 3
                while col <= 6 and row[col - 1].fill.start_color.index != '00000000':
                    col += 1
                fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet.cell(row=row[0].row, column=col).fill = fill
                workbook.save(EXCEL_FILE_PATH)
                break
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def get_players():
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook['Sarga-piros lapok']
    players = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, team = row[0], row[1]
        if name:
            players.append({"name": name, "team": team})

    return players

def update_excel(path, group, team1_name, team1_score, team2_name, team2_score):
    wb = openpyxl.load_workbook(path)
    ws = wb[f"{group} CSOPORT"]

    team1_index = None
    team2_index = None

    for row in range(2, ws.max_row + 1):
        if unidecode(ws.cell(row=row, column=1).value.lower()) == unidecode(team1_name.lower()):
            team1_index = row
        if unidecode(ws.cell(row=row, column=1).value.lower()) == unidecode(team2_name.lower()):
            team2_index = row

    if team1_index is None or team2_index is None:
        print("One or more teams not found in the group.")
        return

    update_team_data(ws, team1_index, team1_score, team2_score)
    update_team_data(ws, team2_index, team2_score, team1_score)

    update_rankings(ws)

    wb.save(path)

def update_team_data(ws, team_index, goals_scored, goals_conceded):
    ws.cell(row=team_index, column=2).value = (ws.cell(row=team_index, column=2).value or 0) + 1

    if goals_scored > goals_conceded:
        ws.cell(row=team_index, column=3).value = (ws.cell(row=team_index, column=3).value or 0) + 1
        ws.cell(row=team_index, column=8).value = (ws.cell(row=team_index, column=8).value or 0) + 3
        ws.cell(row=team_index, column=9).value = add_forma(ws.cell(row=team_index, column=9).value, 'GY')
    elif goals_scored == goals_conceded:
        ws.cell(row=team_index, column=4).value = (ws.cell(row=team_index, column=4).value or 0) + 1
        ws.cell(row=team_index, column=8).value = (ws.cell(row=team_index, column=8).value or 0) + 1
        ws.cell(row=team_index, column=9).value = add_forma(ws.cell(row=team_index, column=9).value, 'D')
    else:
        ws.cell(row=team_index, column=5).value = (ws.cell(row=team_index, column=5).value or 0) + 1
        ws.cell(row=team_index, column=9).value = add_forma(ws.cell(row=team_index, column=9).value, 'V')

    ws.cell(row=team_index, column=6).value = (ws.cell(row=team_index, column=6).value or 0) + goals_scored
    ws.cell(row=team_index, column=7).value = (ws.cell(row=team_index, column=7).value or 0) + (goals_scored - goals_conceded)

def update_rankings(ws):
    rankings = []
    for row in range(2, ws.max_row + 1):
        points = ws.cell(row=row, column=8).value or 0
        goals_diff = ws.cell(row=row, column=7).value or 0
        rankings.append((points, goals_diff, row))

    rankings.sort(key=lambda x: (-x[0], -x[1]))

    for rank, (_, _, row) in enumerate(rankings, start=1):
        ws.cell(row=row, column=10).value = rank

def add_forma(current_forma, result):
    if current_forma is None:
        return result
    else:
        return current_forma + ',' + result

if __name__ == '__main__':
    app.run(debug=True)
