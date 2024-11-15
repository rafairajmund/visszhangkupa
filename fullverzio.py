from openpyxl import load_workbook

# Előre megadott csapatok és celláik
csoport_adatok = {
    "Tiffosi 2010": ("B2", "C2", "D2", "E2", "F2", "G2", "H2", "I2", "J2"),
    "Kalotaszesz": ("B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", "J3"),
    "Kárahegy": ("B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4", "J4"),
    "Prokomisz": ("B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5"),
    "Mad Dogs": ("B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6", "J6"),
    "Screwbolt": ("B7", "C7", "D7", "E7", "F7", "G7", "H7", "I7", "J7"),
    "Fc Finnviz": ("B8", "C8", "D8", "E8", "F8", "G8", "H8", "I8", "J8")
}

# Excel fájl betöltése
wb = load_workbook('D:/visszhang/EREDMENYEK.xlsx')
ws = wb.active

# Segédfüggvény egy cella értékének növeléséhez
def increment_cell_value(cell, value=1):
    if ws[cell].value is None:
        ws[cell].value = value
    else:
        ws[cell].value += value

# Csapatok neveinek bekérése
csapat1_nev = input("Add meg az első csapat nevét: ")
csapat2_nev = input("Add meg a második csapat nevét: ")

# Csapatok celláinak lekérése
csapat1 = csoport_adatok.get(csapat1_nev)
csapat2 = csoport_adatok.get(csapat2_nev)

if not csapat1 or not csapat2:
    print("Hiba: Egyik vagy mindkét csapat neve nincs a listában.")
else:
    # Gólok bekérése
    csapat1_gol = int(input(f"Add meg {csapat1_nev} által rúgott gólok számát: "))
    csapat2_gol = int(input(f"Add meg {csapat2_nev} által rúgott gólok számát: "))

    # Lejátszott meccsek frissítése
    increment_cell_value(csapat1[0], 1)
    increment_cell_value(csapat2[0], 1)

    # Győzelem, döntetlen, vereség frissítése
    if csapat1_gol > csapat2_gol:  # Csapat1 nyert
        increment_cell_value(csapat1[1], 1)
        increment_cell_value(csapat2[3], 1)
        increment_cell_value(csapat1[6], 3)
        ws[csapat1[7]].value = "GY"
        ws[csapat2[7]].value = "V"
    elif csapat1_gol < csapat2_gol:  # Csapat2 nyert
        increment_cell_value(csapat2[1], 1)
        increment_cell_value(csapat1[3], 1)
        increment_cell_value(csapat2[6], 3)
        ws[csapat2[7]].value = "GY"
        ws[csapat1[7]].value = "V"
    else:  # Döntetlen
        increment_cell_value(csapat1[2], 1)
        increment_cell_value(csapat2[2], 1)
        increment_cell_value(csapat1[6], 1)
        increment_cell_value(csapat2[6], 1)
        ws[csapat1[7]].value = "D"
        ws[csapat2[7]].value = "D"

    # Gólok és gólkülönbség frissítése
    increment_cell_value(csapat1[4], csapat1_gol)
    increment_cell_value(csapat2[4], csapat2_gol)
    ws[csapat1[5]].value = (ws[csapat1[4]].value or 0) - csapat2_gol
    ws[csapat2[5]].value = (ws[csapat2[4]].value or 0) - csapat1_gol

    # Helyezések frissítése a pontok alapján
    csapatok = sorted(
        csoport_adatok.keys(),
        key=lambda x: (ws[csoport_adatok[x][6]].value or 0, ws[csoport_adatok[x][5]].value or 0),
        reverse=True
    )
    for i, csapat in enumerate(csapatok, start=1):
        ws[csoport_adatok[csapat][8]].value = i

    # Módosítások mentése
    wb.save("D:/visszhang/EREDMENYEK.xlsx")
    print("A meccs adatai sikeresen frissítve lettek az Excel fájlban.")
